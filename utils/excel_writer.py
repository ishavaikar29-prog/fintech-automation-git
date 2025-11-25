from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from .error_handler import log_exception, log_info
import os

def style_header(row):
    header_font = Font(bold=True, color="000000")
    header_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
    for cell in row:
        cell.font = header_font
        cell.fill = header_fill

def autosize_columns(ws):
    for idx, col in enumerate(ws.columns, 1):
        max_len = 0
        for cell in col:
            try:
                max_len = max(max_len, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[get_column_letter(idx)].width = max_len + 2

def write_sheet(ws, data, mapping):
    headers = list(mapping.keys())
    ws.append(headers)
    style_header(ws[1])

    for item in data:
        ws.append([item.get(field) for field in mapping.values()])

    autosize_columns(ws)

def create_excel(transaction_summary, scorecard, daily_report, filename="daily_report.xlsx"):
    try:
        log_info("Creating Excel file...")
        wb = Workbook()

        ws1 = wb.active
        ws1.title = "TransactionSummary"
        write_sheet(ws1, transaction_summary, {
            "ID": "id",
            "Amount": "amount",
            "Date": "date",
            "Status": "status",
        })

        ws2 = wb.create_sheet("Scorecard")
        write_sheet(ws2, scorecard, {
            "CustomerID": "customer_id",
            "Score": "score",
            "Month": "month",
        })

        ws3 = wb.create_sheet("DailyReport")
        write_sheet(ws3, daily_report, {
            "ReportID": "id",
            "Title": "title",
            "CreatedAt": "created_at",
        })

        wb.save(filename)
        log_info(f"Saved Excel: {filename}")
        return os.path.abspath(filename)
    except Exception as e:
        log_exception("Excel creation failed", e)
        raise
